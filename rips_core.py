"""
Logica de negocio para el generador de plantilla RIPS.
No depende de la interfaz grafica: recibe rutas de archivo y devuelve
resultados / advertencias, y sabe escribir la plantilla oficial sin
tocar su estructura (solo llena celdas de las hojas transaccion,
usuarios y consultas).
"""
import io
import os
import random
import re
import time
from copy import copy
from datetime import date, datetime

import openpyxl
import requests
from openpyxl.utils import get_column_letter

try:
    import xlrd
except ImportError:  # xlrd solo hace falta si el excel de pagos es .xls
    xlrd = None


_ENCODINGS = ("utf-8-sig", "cp1252", "latin-1")


class RipsError(Exception):
    """Error de negocio que se debe mostrar tal cual al usuario."""


# --------------------------------------------------------------------------
# Lectura de los TXT planos
# --------------------------------------------------------------------------

def _read_text(path):
    raw = open(path, "rb").read()
    for enc in _ENCODINGS:
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _split_lines(path):
    text = _read_text(path)
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if line:
            lines.append(line.split(","))
    return lines


def parse_af0(path):
    rows = _split_lines(path)
    if not rows:
        raise RipsError("El archivo AF0 esta vacio.")
    f = rows[0]
    return {
        "codPrestador": f[0].strip(),
        "nomPrestador": f[1].strip() if len(f) > 1 else "",
        "tipoIdObligado": f[2].strip() if len(f) > 2 else "",
        "numDocumentoIdObligado": f[3].strip() if len(f) > 3 else "",
        "numFactura": f[4].strip() if len(f) > 4 else "",
        "fechaExpedicion": f[5].strip() if len(f) > 5 else "",
        "valorTotalCopago": f[13].strip() if len(f) > 13 else "",
        "valorTotal": f[-1].strip() if f else "",
    }


def parse_us0(path):
    rows = _split_lines(path)
    usuarios = []
    for f in rows:
        if len(f) < 14:
            f = f + [""] * (14 - len(f))
        usuarios.append({
            "tipoDocumentoIdentificacion": f[0].strip(),
            "numDocumentoIdentificacion": f[1].strip(),
            "tipoUsuario": f[3].strip(),
            "apellido1": f[4].strip(),
            "apellido2": f[5].strip(),
            "nombre1": f[6].strip(),
            "nombre2": f[7].strip(),
            "edad": f[8].strip(),
            "unidadMedidaEdad": f[9].strip(),
            "sexo": f[10].strip(),
            "codDepartamento": f[11].strip(),
            "codMunicipio": f[12].strip(),
            "zona": f[13].strip().upper(),
        })
    return usuarios


def parse_ap0(path):
    rows = _split_lines(path)
    procedimientos = []
    for f in rows:
        if len(f) < 15:
            f = f + [""] * (15 - len(f))
        procedimientos.append({
            "numFactura": f[0].strip(),
            "codPrestador": f[1].strip(),
            "tipoDocumentoIdentificacion": f[2].strip(),
            "numDocumentoIdentificacion": f[3].strip(),
            "fechaInicioAtencion": f[4].strip(),
            "numAutorizacion": f[5].strip(),
            "codConsulta": f[6].strip(),
            "viaIngreso": f[7].strip(),
            "modalidadGrupoServicioTecSal": f[8].strip(),
            "vrServicio": f[14].strip(),
        })
    return procedimientos


def parse_ad0(path):
    """En esta compania el AD0 trae UNA linea por cada visita/orden del
    paciente (no un solo total agregado para toda la factura), en el mismo
    orden en que aparecen los bloques de paciente en el AP0."""
    rows = _split_lines(path)
    if not rows:
        raise RipsError("El archivo AD0 esta vacio.")
    registros = []
    for f in rows:
        if len(f) < 6:
            f = f + [""] * (6 - len(f))
        registros.append({
            "numFactura": f[0].strip(),
            "codPrestador": f[1].strip(),
            "cantidad": f[3].strip(),
            "valorCopago": f[4].strip(),
            "valorTotal": f[5].strip(),
        })
    return registros


def parse_ct0(path):
    """Devuelve {nombreArchivo: cantidadRegistros} para validar cuadre. No es obligatorio."""
    rows = _split_lines(path)
    control = {}
    for f in rows:
        if len(f) < 4:
            continue
        nombre = f[2].strip().upper()
        try:
            cantidad = int(f[3].strip())
        except ValueError:
            cantidad = None
        control[nombre] = cantidad
    return control


# --------------------------------------------------------------------------
# Excel "REPORTE ORDENES POR EMPRESA" (opcional) -> copago por visita/orden
# --------------------------------------------------------------------------

def _norm_code(value):
    """Normaliza un valor de codigo (puede llegar como texto o como float de excel)."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def load_reporte_ordenes(path):
    """Lee el Excel 'REPORTE ORDENES POR EMPRESA': un reporte agrupado por
    orden (no es una tabla plana). Cada orden trae:
      - una fila de encabezado, con el paciente en la columna C como
        "TIPO NUMERO" (ej. "CE 5300399" o "CC 36171096"),
      - varias filas de detalle (columna B = codigo CUPS, columna D = valor),
      - una fila final "Total Orden" que trae el copago de esa orden en la
        columna H.
    Devuelve una lista de {"numDocumentoIdentificacion": str, "copago": float}
    en el mismo orden en que aparecen las ordenes en el archivo (una entrada
    por orden/visita)."""
    if path.lower().endswith(".xls"):
        rows = _load_xls_rows(path)
    else:
        rows = _load_xlsx_rows(path)

    visitas = []
    actual = None
    for row in rows[1:]:
        col_a = row[0] if len(row) > 0 else None
        col_c = row[2] if len(row) > 2 else None
        col_h = row[7] if len(row) > 7 else None

        if col_a not in (None, "") and isinstance(col_c, str) and col_c.strip() and " " in col_c.strip():
            doc = col_c.strip().split()[-1]
            actual = {"numDocumentoIdentificacion": _norm_code(doc), "copago": 0.0}
        elif isinstance(col_c, str) and col_c.strip().lower() == "total orden" and actual is not None:
            try:
                actual["copago"] = float(col_h) if col_h not in (None, "") else 0.0
            except (TypeError, ValueError):
                actual["copago"] = 0.0
            visitas.append(actual)
            actual = None
    return visitas


def _load_xlsx_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Estadistica"] if "Estadistica" in wb.sheetnames else wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _load_xls_rows(path):
    if xlrd is None:
        raise RipsError("Falta el paquete 'xlrd' para leer archivos .xls.")
    wb = xlrd.open_workbook(path)
    names = wb.sheet_names()
    sheet_name = "Estadistica" if "Estadistica" in names else names[0]
    sh = wb.sheet_by_name(sheet_name)
    return [sh.row_values(r) for r in range(sh.nrows)]


# --------------------------------------------------------------------------
# Reglas de negocio: fechaNacimiento aleatoria, zona, consecutivos, etc.
# --------------------------------------------------------------------------

ZONA_MAP = {"U": "01", "R": "02"}


def format_fecha_hora(fecha_ddmmyyyy):
    """Convierte 'dd/mm/yyyy' a 'yyyy-mm-dd 00:00' (la hora siempre es 00:00)."""
    fecha_ddmmyyyy = (fecha_ddmmyyyy or "").strip()
    if not fecha_ddmmyyyy:
        return ""
    try:
        d = datetime.strptime(fecha_ddmmyyyy, "%d/%m/%Y")
    except ValueError:
        return fecha_ddmmyyyy
    return d.strftime("%Y-%m-%d") + " 00:00"


def random_birthdate_for_age(age, today=None):
    """Genera una fecha de nacimiento aleatoria tal que, comparada con 'today',
    de exactamente 'age' anhos cumplidos."""
    today = today or date.today()
    month = random.randint(1, today.month)
    if month == today.month:
        day = random.randint(1, max(1, today.day))
    else:
        day = random.randint(1, 28)
    year = today.year - age
    return date(year, month, day)


def build_fecha_nacimiento(edad_str, today=None):
    today = today or date.today()
    try:
        edad = int(float(edad_str))
    except (TypeError, ValueError):
        edad = 0
    if edad <= 18:
        target_age = random.randint(19, 65)
    else:
        target_age = edad
    return random_birthdate_for_age(target_age, today)


# El NIT del obligado (num_DocumentoIdObligado) es siempre este valor fijo,
# sin importar lo que traiga el AF0 (evita errores de "no existe la empresa"
# cuando el AF0 llega con el NIT incompleto, sin el digito de verificacion).
NIT_OBLIGADO = "800091457"


def build_transaccion(af0, num_factura):
    return {
        "numDocumentoIdObligado": NIT_OBLIGADO,
        "numFactura": num_factura,
        "tipoNota": "NA",
        "numNota": "",
    }


TIPO_USUARIO_OPCIONES = {
    "Regimen Especial": "06",
    "Empresa Privada": "11",
}


def build_usuarios(us0_rows, num_obligado, tipo_usuario_fijo, today=None):
    def sort_key(u):
        try:
            return int(u["numDocumentoIdentificacion"])
        except ValueError:
            return u["numDocumentoIdentificacion"]

    ordenados = sorted(us0_rows, key=sort_key)
    usuarios = []
    warnings = []
    for i, u in enumerate(ordenados, start=1):
        zona_raw = u["zona"]
        zona = ZONA_MAP.get(zona_raw, "")
        if not zona:
            warnings.append(
                f"Usuario {u['numDocumentoIdentificacion']}: zona '{zona_raw}' desconocida, se dejo vacia."
            )
        fecha_nac = build_fecha_nacimiento(u["edad"], today)
        municipio = f"{u['codDepartamento'].zfill(2)}{u['codMunicipio'].zfill(3)}"
        usuarios.append({
            "tipoDocumentoIdentificacion": "CC",
            "numDocumentoIdentificacion": u["numDocumentoIdentificacion"],
            "num_DocumentoIdObligado": num_obligado,
            "tipoUsuario": tipo_usuario_fijo,
            "fechaNacimiento": fecha_nac,
            "codSexo": u["sexo"],
            "codPaisResidencia": "170",
            "codMunicipioResidencia": municipio,
            "codZonaTerritorialResidencia": zona,
            "incapacidad": "NO",
            "codPaisOrigen": "170",
            "registroSIRAS": "",
            "consecutivo": i,
        })
    return usuarios, warnings


def _agrupar_ap0_por_visita(ap0_rows):
    """Agrupa las filas del AP0 en 'visitas': bloques contiguos del mismo
    paciente, en el mismo orden en que aparecen en el archivo (una visita =
    una orden = una linea del AD0 = una orden del reporte de ordenes)."""
    grupos = []
    doc_actual = None
    bloque = None
    for p in ap0_rows:
        doc = p["numDocumentoIdentificacion"]
        if doc != doc_actual or bloque is None:
            bloque = []
            grupos.append((doc, bloque))
            doc_actual = doc
        bloque.append(p)
    return grupos


def _repartir_copago(lineas_con_valor, copago):
    """Reparte 'copago' entre las lineas de una visita: se ordenan de mayor a
    menor vrServicio y se va asignando min(restante, vrServicio) a cada una
    hasta cubrirlo. Si una sola linea alcanza a cubrir el copago, se le
    asigna todo a esa (la de mayor valor). Nunca deja una linea con
    vrServicio - valorPagoModerador negativo."""
    pagos = {}
    restante = copago
    for p in sorted(lineas_con_valor, key=lambda p: p["_vr"], reverse=True):
        if restante <= 0:
            pagos[id(p)] = 0.0
            continue
        asignado = min(restante, p["_vr"])
        pagos[id(p)] = asignado
        restante -= asignado
    return pagos


def _doc_sort_key(doc):
    """Mismo criterio de orden que build_usuarios (numerico ascendente), para
    que consecutivoUsuario en 'procedimientos' quede en el mismo orden de
    numDocumentoIdentificacion en 'usuarios'."""
    try:
        return (0, int(doc))
    except ValueError:
        return (1, doc)


def build_procedimientos(ap0_rows, num_obligado, visitas_reporte):
    """Arma las filas de la hoja 'procedimientos'. El copago
    (valorPagoModerador) se saca del Excel 'REPORTE ORDENES POR EMPRESA'
    (una orden = una visita = un bloque contiguo del mismo paciente en el
    AP0) y se reparte entre las lineas de esa visita con _repartir_copago."""
    from collections import defaultdict, deque

    visitas_por_doc = defaultdict(deque)
    for v in visitas_reporte or []:
        visitas_por_doc[_norm_code(v["numDocumentoIdentificacion"])].append(v["copago"])

    grupos = _agrupar_ap0_por_visita(ap0_rows)

    sin_copago_reporte = 0
    filas = []

    for doc, lineas in grupos:
        doc_norm = _norm_code(doc)

        cola = visitas_por_doc.get(doc_norm)
        if cola:
            copago_visita = cola.popleft()
        else:
            copago_visita = 0.0
            if visitas_reporte is not None:
                sin_copago_reporte += 1

        for p in lineas:
            try:
                p["_vr"] = int(float(p["vrServicio"])) if p["vrServicio"] else 0
            except ValueError:
                p["_vr"] = 0

        con_valor = [p for p in lineas if p["_vr"] > 0]
        pagos = _repartir_copago(con_valor, copago_visita)

        for p in lineas:
            if p["_vr"] <= 0:
                continue  # las lineas con vrServicio en 0 no se escriben en la plantilla
            pago = pagos.get(id(p), 0.0)
            concepto_recaudo = "05" if pago == 0 else "03"
            num_fev_pago_moderador = 0 if pago > 0 else ""

            filas.append({
                "num_DocumentoIdObligado": num_obligado,
                "consecutivoUsuario": p["numDocumentoIdentificacion"],
                "codPrestador": "410010011901",
                "fechaInicioAtencion": format_fecha_hora(p["fechaInicioAtencion"]),
                "idMIPRES": "",
                "numAutorizacion": p["numAutorizacion"],
                "codProcedimiento": p["codConsulta"],
                "viaIngresoServicioSalud": "01",
                "modalidadGrupoServicioTecSal": "01",
                "grupoServicios": "02",
                "codServicio": 706,
                "finalidadTecnologiaSalud": "15",
                "tipoDocumentoIdentificacion": "CC",
                "numDocumentoIdentificacion": "36170543",
                "codDiagnosticoPrincipal": "Z017",
                "codDiagnosticoPrincipalCIE11": "Z017",
                "nomCodDiagnosticoPrincipalCIE11": "EXAMEN DE LABORATORIO",
                "codDiagnosticoRelacionado": "",
                "codDiagnosticoRelacionadoCIE11": "",
                "nomCodDiagnosticoRelacionadoCIE11": "",
                "codComplicacion": "",
                "codComplicacionCIE11": "",
                "nomCodComplicacionCIE11": "",
                "vrServicio": p["_vr"],
                "conceptoRecaudo": concepto_recaudo,
                "valorPagoModerador": int(pago) if float(pago).is_integer() else pago,
                "numFEVPagoModerador": num_fev_pago_moderador,
                "codigoVIDA": "",
                "_doc_sort": _doc_sort_key(doc_norm),
                "_doc": doc_norm,
            })

    # Orden final: pacientes en el mismo orden numerico ascendente que la
    # hoja 'usuarios' (build_usuarios); dentro de cada paciente (uniendo
    # todas sus visitas), lineas por codProcedimiento ascendente, y un
    # consecutivo 1..N propio de cada paciente.
    filas.sort(key=lambda f: (f["_doc_sort"], _norm_code(f["codProcedimiento"])))

    procedimientos = []
    last_doc = None
    seq = 0
    for f in filas:
        if f["_doc"] != last_doc:
            seq = 1
            last_doc = f["_doc"]
        else:
            seq += 1
        f["consecutivo"] = seq
        del f["_doc_sort"]
        del f["_doc"]
        procedimientos.append(f)

    return procedimientos, sin_copago_reporte


# --------------------------------------------------------------------------
# Escritura en la plantilla oficial (sin tocar su estructura)
# --------------------------------------------------------------------------

TRANSACCION_COLS = ["numDocumentoIdObligado", "numFactura", "tipoNota", "numNota"]

USUARIOS_COLS = [
    "tipoDocumentoIdentificacion", "numDocumentoIdentificacion", "num_DocumentoIdObligado",
    "tipoUsuario", "fechaNacimiento", "codSexo", "codPaisResidencia", "codMunicipioResidencia",
    "codZonaTerritorialResidencia", "incapacidad", "codPaisOrigen", "registroSIRAS", "consecutivo",
]

PROCEDIMIENTOS_COLS = [
    "num_DocumentoIdObligado", "consecutivoUsuario", "codPrestador", "fechaInicioAtencion",
    "idMIPRES", "numAutorizacion", "codProcedimiento", "viaIngresoServicioSalud",
    "modalidadGrupoServicioTecSal", "grupoServicios", "codServicio", "finalidadTecnologiaSalud",
    "tipoDocumentoIdentificacion", "numDocumentoIdentificacion", "codDiagnosticoPrincipal",
    "codDiagnosticoPrincipalCIE11", "nomCodDiagnosticoPrincipalCIE11", "codDiagnosticoRelacionado",
    "codDiagnosticoRelacionadoCIE11", "nomCodDiagnosticoRelacionadoCIE11", "codComplicacion",
    "codComplicacionCIE11", "nomCodComplicacionCIE11", "vrServicio", "conceptoRecaudo",
    "valorPagoModerador", "numFEVPagoModerador", "codigoVIDA", "consecutivo",
]


def _expand_table_ranges(ws, n_cols, last_row):
    """Las hojas usan Tablas de Excel (ListObjects) con rango fijo (ej. A1:M2).
    Si no se expande ese rango al escribir mas filas, Excel/las herramientas de
    validacion RIPS solo reconocen la primera fila de datos como parte de la
    tabla. Aqui se ajusta el 'ref' de la tabla para cubrir todas las filas
    escritas, sin tocar columnas, encabezados ni ninguna otra propiedad."""
    if not ws.tables:
        return
    last_col_letter = get_column_letter(n_cols)
    new_ref = f"A1:{last_col_letter}{last_row}"
    for table_name in list(ws.tables.keys()):
        ws.tables[table_name].ref = new_ref


def fill_template(template_path, output_path, transaccion, usuarios, procedimientos):
    """Escribe los datos en la plantilla oficial. La fila 2 de cada hoja ya
    trae el formato de celda correcto (numerico vs texto) definido por la
    plantilla; las filas nuevas que abrimos (3 en adelante) no heredan ese
    formato automaticamente, y el conversor de JSON de SISPRO usa ese
    formato -no solo el valor- para decidir si un campo es numero o texto.
    Por eso se copia el estilo de la fila 2 a cada fila nueva, columna por
    columna, para que todas las filas se conviertan igual."""
    wb = openpyxl.load_workbook(template_path, keep_vba=True)

    for sheet_name, cols, rows in (
        ("transaccion", TRANSACCION_COLS, [transaccion]),
        ("usuarios", USUARIOS_COLS, usuarios),
        ("procedimientos", PROCEDIMIENTOS_COLS, procedimientos),
    ):
        ws = wb[sheet_name]
        template_styles = [ws.cell(row=2, column=c)._style for c in range(1, len(cols) + 1)]
        last_row = 1
        for row_idx, data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(cols, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=data.get(col_name))
                if row_idx > 2:
                    cell._style = copy(template_styles[col_idx - 1])
            last_row = row_idx
        _expand_table_ranges(ws, len(cols), max(last_row, 2))

    try:
        wb.save(output_path)
    except PermissionError:
        raise RipsError(
            f"No se pudo guardar porque el archivo esta abierto en Excel:\n{output_path}\n\n"
            "Cierra el archivo y vuelve a intentar."
        )


# --------------------------------------------------------------------------
# Orquestacion completa
# --------------------------------------------------------------------------

def procesar(af0_path, us0_path, ap0_path, ad0_path, template_path, output_path,
             num_factura, tipo_usuario, ct0_path=None, pago_path=None, today=None):
    """Ejecuta todo el flujo. Devuelve (resumen_dict, lista_de_advertencias)."""
    warnings = []

    if not num_factura or not num_factura.strip():
        raise RipsError("Debes ingresar el prefijo y numero de factura.")
    num_factura = num_factura.strip()

    if tipo_usuario not in TIPO_USUARIO_OPCIONES.values():
        raise RipsError("Debes seleccionar el tipo de usuario (Regimen Especial o Empresa Privada).")

    af0 = parse_af0(af0_path)
    us0_rows = parse_us0(us0_path)
    ap0_rows = parse_ap0(ap0_path)
    ad0_rows = parse_ad0(ad0_path)

    if af0.get("numFactura") and af0["numFactura"] != num_factura:
        warnings.append(
            f"El numFactura ingresado ({num_factura}) es distinto al que trae el AF0 ({af0['numFactura']})."
        )

    if ct0_path:
        control = parse_ct0(ct0_path)

        def _cantidad_para(prefijo):
            for nombre, cantidad in control.items():
                if nombre.startswith(prefijo):
                    return cantidad
            return None

        checks = [
            ("AF0", 1),
            ("US0", len(us0_rows)),
            ("AP0", len(ap0_rows)),
            ("AD0", len(ad0_rows)),
        ]
        for prefijo, cantidad_real in checks:
            esperado = _cantidad_para(prefijo)
            if esperado is not None and esperado != cantidad_real:
                warnings.append(
                    f"CT0 esperaba {esperado} registros de {prefijo}, pero se leyeron {cantidad_real}."
                )

    visitas_reporte = None
    if pago_path:
        visitas_reporte = load_reporte_ordenes(pago_path)

    transaccion = build_transaccion(af0, num_factura)
    usuarios, warn_usuarios = build_usuarios(us0_rows, NIT_OBLIGADO, tipo_usuario, today)
    warnings.extend(warn_usuarios)

    procedimientos, sin_copago = build_procedimientos(ap0_rows, NIT_OBLIGADO, visitas_reporte)
    if visitas_reporte is not None and sin_copago:
        warnings.append(
            f"{sin_copago} visita(s) del AP0 no encontraron una orden correspondiente en el "
            "Reporte de ordenes por empresa adjunto (se dejo el copago en 0)."
        )
    if visitas_reporte is None:
        warnings.append(
            "No se adjunto el Excel 'Reporte ordenes por empresa': valorPagoModerador se "
            "dejo en 0 para todos los registros."
        )

    suma_vr_servicio = sum(c["vrServicio"] for c in procedimientos)
    suma_pago_moderador = sum(c["valorPagoModerador"] for c in procedimientos)

    try:
        valor_ad0_total = sum(int(float(r["valorTotal"])) for r in ad0_rows if r["valorTotal"])
    except ValueError:
        valor_ad0_total = None
    cuadre_ok = valor_ad0_total is not None and suma_vr_servicio == valor_ad0_total
    if valor_ad0_total is not None and not cuadre_ok:
        warnings.append(
            f"La suma de vrServicio ({suma_vr_servicio}) no coincide con la suma de valores totales del AD0 ({valor_ad0_total})."
        )

    try:
        valor_ad0_copago = sum(int(float(r["valorCopago"])) for r in ad0_rows if r["valorCopago"])
    except ValueError:
        valor_ad0_copago = None
    if valor_ad0_copago is not None and suma_pago_moderador != valor_ad0_copago:
        warnings.append(
            f"La suma de valorPagoModerador ({suma_pago_moderador}) no coincide con la suma de copagos del AD0 ({valor_ad0_copago})."
        )

    if af0.get("valorTotalCopago"):
        try:
            valor_af0_copago = int(float(af0["valorTotalCopago"]))
            if valor_af0_copago != suma_pago_moderador:
                warnings.append(
                    f"La suma de valorPagoModerador ({suma_pago_moderador}) no coincide con el valorTotalCopago del AF0 ({valor_af0_copago})."
                )
        except ValueError:
            pass

    fill_template(template_path, output_path, transaccion, usuarios, procedimientos)

    resumen = {
        "numFactura": transaccion["numFactura"],
        "numDocumentoIdObligado": transaccion["numDocumentoIdObligado"],
        "totalUsuarios": len(usuarios),
        "totalConsultas": len(procedimientos),
        "sumaVrServicio": suma_vr_servicio,
        "valorTotalAD0": valor_ad0_total,
        "cuadreOk": cuadre_ok,
        "outputPath": output_path,
    }
    return resumen, warnings


# --------------------------------------------------------------------------
# Conversion a JSON usando el conversor oficial de SISPRO (sin navegador)
# --------------------------------------------------------------------------

SISPRO_ORIGIN = "https://fevrips.sispro.gov.co"
SISPRO_BASE = f"{SISPRO_ORIGIN}/SPConvertidorJSON"
SISPRO_UPLOAD_URL = f"{SISPRO_BASE}/Home/UploadFile"
SISPRO_PLANTILLA_URL = f"{SISPRO_BASE}/plantilla/plantilla_fev-rips_v%201_0.xlsm"

_SISPRO_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)


def descargar_plantilla_sispro(output_path, timeout=60):
    """Descarga la plantilla oficial RIPS MINSALUD directamente desde SISPRO
    (el mismo archivo que sirve el enlace 'Descargar Archivo Excel' de
    https://fevrips.sispro.gov.co/SPConvertidorJSON), sin necesidad de
    navegador."""
    try:
        resp = requests.get(
            SISPRO_PLANTILLA_URL, timeout=timeout,
            headers={"User-Agent": _SISPRO_UA, "Referer": SISPRO_BASE},
        )
        resp.raise_for_status()
    except requests.RequestException as exc:
        raise RipsError(f"No se pudo descargar la plantilla desde SISPRO: {exc}")

    try:
        with open(output_path, "wb") as f:
            f.write(resp.content)
    except PermissionError:
        raise RipsError(
            f"No se pudo guardar porque el archivo esta abierto en Excel:\n{output_path}\n\n"
            "Cierra el archivo y vuelve a intentar."
        )
    return output_path


def convertir_json_sispro(xlsm_path, output_json_path=None, intentos=4):
    """Sube la plantilla al conversor oficial de SISPRO (fevrips.sispro.gov.co)
    y descarga el JSON resultante. No usa navegador, solo HTTP directo (el
    mismo endpoint que usa el boton 'Convertir JSON' de la pagina).

    A veces la descarga inmediata da 404 porque el servidor esta detras de
    un balanceador de carga y el nodo que atiende el GET no es el mismo que
    genero el archivo; por eso se reintenta el ciclo completo varias veces.
    """
    if not output_json_path:
        base, _ = os.path.splitext(xlsm_path)
        output_json_path = base + ".json"

    session = requests.Session()
    session.headers.update({"User-Agent": _SISPRO_UA, "Referer": SISPRO_BASE})

    ultimo_error = None
    for intento in range(1, intentos + 1):
        try:
            session.get(SISPRO_BASE, timeout=30)

            with open(xlsm_path, "rb") as f:
                files = {
                    "file": (
                        os.path.basename(xlsm_path), f,
                        "application/vnd.ms-excel.sheet.macroEnabled.12",
                    )
                }
                resp = session.post(SISPRO_UPLOAD_URL, files=files, timeout=120)
            resp.raise_for_status()
            data = resp.json()
            json_url = data.get("jsonFileUrl")
            if not json_url:
                raise RipsError(f"SISPRO no devolvio la URL del JSON: {data}")

            descarga = session.get(SISPRO_ORIGIN + json_url, timeout=60)
            if descarga.status_code == 200 and descarga.content:
                with open(output_json_path, "wb") as out:
                    out.write(descarga.content)
                return output_json_path
            ultimo_error = f"La descarga del JSON devolvio HTTP {descarga.status_code}."
        except requests.RequestException as exc:
            ultimo_error = str(exc)

        time.sleep(1.5)

    raise RipsError(
        f"No se pudo convertir a JSON en SISPRO despues de {intentos} intentos. "
        f"Ultimo error: {ultimo_error}"
    )


# --------------------------------------------------------------------------
# Excel adicional de facturacion (formato "SUBIR" / World Office)
# --------------------------------------------------------------------------

TIPO_DOCUMENTO_FACTURA_OPCIONES = ["FV", "PED"]

SUBIR_HEADERS = [
    "Encabezados : Empresa",
    "Encabezados : Tipo_Documento",
    "Encabezados : Prefijo",
    "Encabezados : DocumentoNúmero",
    "Encabezados : Fecha",
    "Encabezados : TerceroInterno",
    "Encabezados : TerceroExterno",
    "Encabezados : Nota",
    "Encabezados : FormaPago",
    "Encabezados : Fecha Entrega",
    "Encabezados : Prefijo Documento Externo",
    "Encabezados : Número_Documento_Externo",
    "Encabezados : Verificado",
    "Encabezados : Anulado",
    "Encabezados : Personalizado1",
    "Encabezados : Personalizado2",
    "Encabezados : Personalizado3",
    "Encabezados : Personalizado4",
    "Encabezados : Personalizado5",
    "Encabezados : Personalizado6",
    "Encabezados : Personalizado7",
    "Encabezados : Personalizado8",
    "Encabezados : Personalizado9",
    "Encabezados : Personalizado10",
    "Encabezados : Personalizado11",
    "Encabezados : Personalizado12",
    "Encabezados : Personalizado13",
    "Encabezados : Personalizado14",
    "Encabezados : Personalizado15",
    "Encabezados : Sucursal",
    "Encabezados : Clasificación",
    "Movimientos de Inventario : Producto",
    "Movimientos de Inventario : Bodega_",
    "Movimientos de Inventario : UnidadDeMedida_",
    "Movimientos de Inventario : Cantidad__",
    "Movimientos de Inventario : Iva_",
    "Movimientos de Inventario : Valor_Unitario__",
    "Movimientos de Inventario : Descuento_",
    "Movimientos de Inventario : Vencimiento_",
    "Movimientos de Inventario : Nota_",
    "Movimientos de Inventario : Centro_de_Costo",
    "Movimientos de Inventario : Personalizado_1",
    "Movimientos de Inventario : Personalizado_2",
    "Movimientos de Inventario : Personalizado_3",
    "Movimientos de Inventario : Personalizado_4",
    "Movimientos de Inventario : Personalizado_5",
]

EMPRESA_FACTURA = "LABORATORIO DIAGNOSTICAMOS SAS"


def build_factura_plana(ap0_rows, af0, nit_cliente, tipo_documento, prefijo, num_doc, fecha, cedula_trabajador):
    """Arma las filas del Excel de facturacion (formato SUBIR / World Office).
    Las lineas del AP0 se agrupan por (codConsulta, valor unitario): si el
    mismo producto se repite con el mismo valor unitario, se consolida en
    una sola fila sumando la Cantidad, en vez de repetir la misma linea
    muchas veces (Valor_Unitario es el vrServicio de cada linea del AP0,
    que ya viene unitario)."""
    if tipo_documento not in TIPO_DOCUMENTO_FACTURA_OPCIONES:
        raise RipsError("Debes seleccionar el tipo de documento (FV o PED).")
    if not nit_cliente or not nit_cliente.strip().isdigit() or len(nit_cliente.strip()) != 9:
        raise RipsError("El NIT Cliente debe tener exactamente 9 digitos.")
    if not cedula_trabajador or not cedula_trabajador.strip():
        raise RipsError("Debes ingresar la cedula del trabajador.")
    if not prefijo or not num_doc:
        raise RipsError("Debes ingresar el prefijo y el consecutivo de la factura.")

    nit_cliente = nit_cliente.strip()
    cedula_trabajador = cedula_trabajador.strip()
    fecha_texto = fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else str(fecha)

    agrupados = {}
    orden = []
    for p in ap0_rows:
        try:
            valor_unitario = int(float(p["vrServicio"])) if p["vrServicio"] else 0
        except ValueError:
            valor_unitario = 0
        clave = (p["codConsulta"], valor_unitario)
        if clave not in agrupados:
            agrupados[clave] = 0
            orden.append(clave)
        agrupados[clave] += 1

    filas = []
    suma_control = 0
    for codigo, valor_unitario in orden:
        cantidad = agrupados[(codigo, valor_unitario)]
        suma_control += cantidad * valor_unitario

        fila = {h: None for h in SUBIR_HEADERS}
        fila["Encabezados : Empresa"] = EMPRESA_FACTURA
        fila["Encabezados : Tipo_Documento"] = tipo_documento
        fila["Encabezados : Prefijo"] = prefijo
        fila["Encabezados : DocumentoNúmero"] = num_doc
        fila["Encabezados : Fecha"] = fecha_texto
        fila["Encabezados : TerceroInterno"] = cedula_trabajador
        fila["Encabezados : TerceroExterno"] = nit_cliente
        fila["Encabezados : Nota"] = "FACTURA DE VENTA"
        fila["Encabezados : FormaPago"] = "Credito"
        fila["Encabezados : Fecha Entrega"] = fecha_texto
        fila["Encabezados : Verificado"] = 0
        fila["Encabezados : Anulado"] = 0
        fila["Movimientos de Inventario : Producto"] = codigo
        fila["Movimientos de Inventario : Bodega_"] = "Principal"
        fila["Movimientos de Inventario : UnidadDeMedida_"] = "Und."
        fila["Movimientos de Inventario : Cantidad__"] = cantidad
        fila["Movimientos de Inventario : Iva_"] = 0
        fila["Movimientos de Inventario : Valor_Unitario__"] = valor_unitario
        fila["Movimientos de Inventario : Descuento_"] = 0
        fila["Movimientos de Inventario : Vencimiento_"] = fecha_texto
        filas.append(fila)

    cuadre_ok = None
    valor_af0 = None
    if af0.get("valorTotal"):
        try:
            valor_af0 = int(float(af0["valorTotal"]))
            cuadre_ok = valor_af0 == suma_control
        except ValueError:
            valor_af0 = None

    return filas, suma_control, valor_af0, cuadre_ok


def guardar_factura_plana(filas, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUBIR"
    ws.append(SUBIR_HEADERS)
    for fila in filas:
        ws.append([fila.get(h) for h in SUBIR_HEADERS])

    try:
        wb.save(output_path)
    except PermissionError:
        raise RipsError(
            f"No se pudo guardar porque el archivo esta abierto en Excel:\n{output_path}\n\n"
            "Cierra el archivo y vuelve a intentar."
        )


# --------------------------------------------------------------------------
# Excel de creacion de usuarios (otro sistema, no es el RIPS)
# --------------------------------------------------------------------------

CIUDAD_FIJA = "Neiva"
DIRECCION_FIJA = "Carrera 11 No 7 – 45"
TELEFONO_FIJO = "3100000000"
EMAIL_FIJO = "noinforma@gmail.com"

SEXO_MAP = {"M": "Masculino", "F": "Femenino"}

CREACION_USUARIOS_HEADERS = [
    "Tipo identificación", "Identificación", "Ciudad", "Primer Nombre", "Segundo Nombre",
    "Primer Apellido", "Segundo Apellido", "Fecha de Nacimiento", "Ciudad de Nacimiento",
    "Sexo", "Ciudad de Residencia", "Dirección", "Teléfono", "Email", "Tipo de Usuario",
    "", "", "Entidad",
]


def build_excel_usuarios(us0_rows, tipo_usuario_fijo, entidad_nit, today=None):
    """Arma las filas del Excel de creacion de usuarios (para otro sistema, no
    es el RIPS). Reutiliza la misma regla de fechaNacimiento aleatoria y el
    mismo tipoUsuario fijo (desplegable) que la hoja 'usuarios' de la plantilla RIPS."""
    if not entidad_nit or not entidad_nit.strip().isdigit() or len(entidad_nit.strip()) != 9:
        raise RipsError("La Entidad (NIT) debe tener exactamente 9 digitos.")
    if tipo_usuario_fijo not in TIPO_USUARIO_OPCIONES.values():
        raise RipsError("Debes seleccionar el tipo de usuario (Regimen Especial o Empresa Privada).")

    entidad_nit = entidad_nit.strip()

    def sort_key(u):
        try:
            return int(u["numDocumentoIdentificacion"])
        except ValueError:
            return u["numDocumentoIdentificacion"]

    ordenados = sorted(us0_rows, key=sort_key)
    warnings = []
    filas = []
    for u in ordenados:
        sexo_raw = u["sexo"].strip().upper()
        sexo = SEXO_MAP.get(sexo_raw, "")
        if not sexo:
            warnings.append(f"Usuario {u['numDocumentoIdentificacion']}: sexo '{sexo_raw}' desconocido, se dejo vacio.")

        fecha_nac = build_fecha_nacimiento(u["edad"], today)

        filas.append({
            "Tipo identificación": "CC",
            "Identificación": u["numDocumentoIdentificacion"],
            "Ciudad": CIUDAD_FIJA,
            "Primer Nombre": u["nombre1"],
            "Segundo Nombre": u["nombre2"],
            "Primer Apellido": u["apellido1"],
            "Segundo Apellido": u["apellido2"],
            "Fecha de Nacimiento": fecha_nac,
            "Ciudad de Nacimiento": CIUDAD_FIJA,
            "Sexo": sexo,
            "Ciudad de Residencia": CIUDAD_FIJA,
            "Dirección": DIRECCION_FIJA,
            "Teléfono": TELEFONO_FIJO,
            "Email": EMAIL_FIJO,
            "Tipo de Usuario": tipo_usuario_fijo,
            "Entidad": entidad_nit,
        })

    return filas, warnings


def guardar_excel_usuarios(filas, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(CREACION_USUARIOS_HEADERS)
    for fila in filas:
        ws.append([fila.get(h) for h in CREACION_USUARIOS_HEADERS])

    try:
        wb.save(output_path)
    except PermissionError:
        raise RipsError(
            f"No se pudo guardar porque el archivo esta abierto en Excel:\n{output_path}\n\n"
            "Cierra el archivo y vuelve a intentar."
        )


# --------------------------------------------------------------------------
# Ajuste del XML de la factura electronica segun el tipoUsuario de la hoja
# 'usuarios' de la plantilla RIPS (06 = Regimen Especial, 11 = Empresa Privada)
# --------------------------------------------------------------------------

# NOTA: los valores exactos de MODALIDAD_PAGO y COBERTURA_PLAN_BENEFICIOS por
# tipoUsuario los definio el usuario. La ubicacion/estructura de estos campos
# dentro del XML (nombres de etiqueta, namespaces) se resuelve buscando por
# nombre de etiqueta local (ignorando el prefijo del namespace, ej. cac:,
# ext:, etc.) para que funcione sin importar el namespace exacto del XML real.
MODALIDAD_PAGO_POR_TIPO = {
    "11": {"schemeID": "04", "schemeName": "salud_modalidad_pago.gc", "texto": "Pago por evento"},
    "06": {"schemeID": "04", "schemeName": "salud_modalidad_pago.gc", "texto": "Pago por evento"},
}

COBERTURA_PLAN_POR_TIPO = {
    "11": {"schemeID": "11", "schemeName": "salud_cobertuta.gc", "texto": "Plan medicina prepagada"},
    "06": {"schemeID": "13", "schemeName": "salud_cobertura.gc", "texto": "Cobertura Régimen Especial o Excepción"},
}


def _set_attr(attrs_str, name, value):
    """Reemplaza (o agrega si no existe) un atributo dentro del texto crudo
    de atributos de una etiqueta XML, sin tocar los demas atributos."""
    patron_attr = re.compile(re.escape(name) + r'\s*=\s*"[^"]*"')
    nuevo = f'{name}="{value}"'
    if patron_attr.search(attrs_str):
        return patron_attr.sub(nuevo, attrs_str, count=1)
    return attrs_str.rstrip() + f' {nuevo}'


def _reemplazar_additional_info(xml_text, nombre, scheme_id, scheme_name, valor_texto):
    """Busca el bloque '<...Name>NOMBRE</...Name> <...Value ...>...</...Value>'
    (sin importar el prefijo de namespace de las etiquetas) y reemplaza los
    atributos schemeID/schemeName y el texto del Value. Se edita el XML como
    texto (no se reconstruye el arbol) para no alterar namespaces, formato
    ni nada mas del documento original. Devuelve (texto_nuevo, cantidad_reemplazos)."""
    patron = re.compile(
        r'(<(?P<ntag>[A-Za-z0-9_.:-]*Name)\b[^>]*>\s*' + re.escape(nombre) + r'\s*</(?P=ntag)>\s*'
        r'<(?P<vtag>[A-Za-z0-9_.:-]*Value)\b)([^>]*)>(.*?)</(?P=vtag)>',
        re.DOTALL,
    )

    def _sub(m):
        inicio = m.group(1)
        vtag = m.group("vtag")
        attrs = _set_attr(m.group(4), "schemeID", scheme_id)
        attrs = _set_attr(attrs, "schemeName", scheme_name)
        return f'{inicio}{attrs}>{valor_texto}</{vtag}>'

    return patron.subn(_sub, xml_text, count=1)


def ajustar_xml_factura(xml_path, tipo_usuario, output_path):
    """Ajusta en el XML de la factura electronica los campos MODALIDAD_PAGO
    y COBERTURA_PLAN_BENEFICIOS segun el tipoUsuario (06 u 11) de la hoja
    'usuarios' de la plantilla RIPS."""
    if tipo_usuario not in MODALIDAD_PAGO_POR_TIPO:
        raise RipsError("Tipo de usuario no reconocido para ajustar el XML (debe ser 06 u 11).")

    raw = open(xml_path, "rb").read()
    texto = None
    for enc in _ENCODINGS + ("utf-8",):
        try:
            texto = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise RipsError("No se pudo leer la codificacion del XML.")

    datos_modalidad = MODALIDAD_PAGO_POR_TIPO[tipo_usuario]
    datos_cobertura = COBERTURA_PLAN_POR_TIPO[tipo_usuario]

    texto, n1 = _reemplazar_additional_info(
        texto, "MODALIDAD_PAGO",
        datos_modalidad["schemeID"], datos_modalidad["schemeName"], datos_modalidad["texto"],
    )
    texto, n2 = _reemplazar_additional_info(
        texto, "COBERTURA_PLAN_BENEFICIOS",
        datos_cobertura["schemeID"], datos_cobertura["schemeName"], datos_cobertura["texto"],
    )

    faltantes = []
    if n1 == 0:
        faltantes.append("MODALIDAD_PAGO")
    if n2 == 0:
        faltantes.append("COBERTURA_PLAN_BENEFICIOS")
    if faltantes:
        raise RipsError(
            "No se encontraron en el XML los campos " + ", ".join(faltantes) + ". "
            "Revisa que sea el XML correcto de la factura electronica."
        )

    # newline="" evita que Python traduzca los saltos de linea del XML
    # original a los del sistema operativo (en Windows duplicaria los \r\n
    # existentes); se preserva el archivo tal cual, solo con el contenido
    # de los campos ajustados.
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        f.write(texto)
    return n1 + n2


def carpeta_salida_rips(base_dir, prefijo, num_factura):
    """Crea (si no existe) y devuelve la carpeta 'RIPS_{prefijo}_{numFactura}'
    donde deben quedar el JSON generado y el XML ajustado de una factura."""
    nombre = f"RIPS_{prefijo}_{num_factura}"
    carpeta = os.path.join(base_dir, nombre)
    os.makedirs(carpeta, exist_ok=True)
    return carpeta
